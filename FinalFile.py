from openpyxl import load_workbook
import  pandas as pd
from datetime import  date
class AccountingCycle:
    def __init__(self,filename,filename2):
        self.filename=filename
        self.filename2=filename2
        self.workbook=load_workbook(self.filename)
        self.sheet=self.workbook.active
        self.workbook1 = load_workbook(self.filename2)
        self.sheet1 = self.workbook1.active
        self.AN=[]
        self.exceldata=[]
        self.exceldata1 = []
        self.Uaccnames=[]
        self.Uaccnames1 = []
        self.ledbal = {}
        self.ledbal1 = {}
        self.NetIncome =[0]
        self.Ownersequity = [0]
        self.ISBal = 0
        self.Assets=["Cash","Supplies","Accounts Receivable","Prepaid Insurance","Equipment","Land"]
        self.Liablities=["Accounts Payable","Notes Payable","Utilities Payable","Interest Payable","Accumulated Depreciation Equipment","Salaries & Wages Payable","Unearned Rent Revenue","Rent Payable","Unearned Service Revenue"]
        self.Capital=["Capital","Investment"]
        self.Drawings=["Owner's Drawing"]
        self.Revenue=["Service Revenue","Rent Revenue"]
        self.Expenses=["Supplies Expense","Salaries & Wages Expense","Insurance Expense","Utilities Expense","Advertising Expense","Interest Expense","Rent Expense","Depreciation Expense"]
        self.IncomeSummary = ["Income Summary"]
        self.read =pd.DataFrame()
        self.read1 = pd.DataFrame()
        self.Adled=pd.DataFrame()
        self.data=pd.DataFrame()
    def Data(self):
        read = pd.read_excel(self.filename)
        self.read=self.read.append(read,ignore_index=True)
        read1 = pd.read_excel(self.filename2)
        self.read1=self.read1.append(read1,ignore_index=True)
        self.Adled=self.Adled.append(read,ignore_index=True)
        self.Adled = self.Adled.append(read1, ignore_index=True)
        self.data=self.data.append(self.read)
        self.data = self.data.append(self.read1,ignore_index=True)
        return self.data

    def GeneralEnteries(self):
        print("--------------------------General Enteries----------------------------------")
        print(self.read)
        sumdr=self.read['Debit'].sum()
        sumcr=self.read['Credit'].sum()
        print("___________________________________________________________________________")
        print(format(sumdr,">63"),format(sumcr,">7"))
        print("___________________________________________________________________________")

    def Tledger(self, AccountName):# for balances of ledger to be enter in Trial Balance

        Dr = self.read.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
        Cr = self.read.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
        if Dr > Cr:
            bal = (Dr) - (Cr)
            self.ledbal1[AccountName] = bal

        if Cr > Dr:
            bal = (Cr) - (Dr)
            self.ledbal1[AccountName] = bal

    def ledger(self,AccountName):#this function wiill print  specific ledger of general enteries
        print("\n")
        print("----------------Ledger of ",AccountName,"------------------")
        print("_____________________________________________________________")
        l = self.read.loc[self.read['Account Title and Explanation'] == AccountName].groupby(['Date', "Debit"]).sum().reset_index()
        Dr=self.read.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
        Cr=self.read.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
        if Dr > Cr :
            print(l)
            bal = (Dr) - (Cr)
            self.ledbal1[AccountName] = bal
            print("_____________________________________________________________")
            print(format(("Bal:"+str(bal)),">21"))
            print("_____________________________________________________________")
        if Cr >Dr :
            print(l)
            bal = (Cr )- (Dr)
            self.ledbal1[AccountName] = bal

            print("_____________________________________________________________")
            print(format(("Bal:"+str(bal)),">27"))
            print("_____________________________________________________________")

    def AccountNames(self):#this function will append all account names in Uaccnames
        for col in self.sheet.iter_cols(min_col=2, max_col=5, min_row=2, values_only=True):
            self.exceldata.append(col)
        accnames = list(self.exceldata[1])
        accnames = list(dict.fromkeys(accnames))
        for i in accnames:
            self.Uaccnames.append(i)
        return accnames
    def Allledger(self):#this function will print all ledger of  general entry
        listAccNames=self.AccountNames()
        for i in listAccNames:
            self.ledger(i)

    def TAllLedger(self):#this funtion will print all ledger balances in Trial Balance
        listAccNames = self.AccountNames()
        for i in listAccNames:
            self.Tledger(i)
    def TrialBalance(self):
        self.TAllLedger()
        print("------------------Trial  Balance------------------------")
        print("-------------------Company Name-------------------------")
        print("----------------For the Month (date)---------------------")
        print("_______________________________________________________")
        print("          Accounts              Debit      Credit")
        print("________________________________________________________")
        cr=[]
        dr=[]
        result1 = list(self.read['Account Title and Explanation'])

        for i in (self.Assets):
            if i in  result1:

                if self.ledbal1[i]>0:
                    dr.append(self.ledbal1[i])
                    print(format(i,">22s"),format(str(self.ledbal1[i]),">14s"))
                else:
                    print(format(i,">22s"), format(str(self.ledbal1[i]),">20s"))
                    cr.append(self.ledbal1[i])

        for i in (self.Liablities):
            if i in  result1:
                if self.ledbal1[i]>0:
                    cr.append(self.ledbal1[i])
                    print(format(i,">22s"), format(str(self.ledbal1[i]),">25s"))
                else:
                    dr.append(self.ledbal1[i])
                    print(format(i,">22s"), format(str(self.ledbal1[i]),">14s"))


        for i in (self.Capital):
            if i in  result1:
                if self.ledbal1[i]>0:
                    cr.append(self.ledbal1[i])
                    print(format(i,">22s"), format(str(self.ledbal1[i]),">25s"))
                else:
                    dr.append(self.ledbal1[i])
                    print(format(i,">22s"), format(str(self.ledbal1[i]),">14s"))

        for i in (self.Drawings):
            if i in  result1:
                if self.ledbal1[i]>0:
                    dr.append(self.ledbal1[i])
                    print(format(i,">22s"), format(str(self.ledbal1[i]),">14s"))
                else:
                    cr.append(self.ledbal1[i])
                    print(format(i,">22s"), format(str(self.ledbal1[i]),">25s"))

        for i in (self.Revenue):
            if i in  result1:
                if self.ledbal1[i]>0:
                    cr.append(self.ledbal1[i])
                    print(format(i,">22s"), format(str(self.ledbal1[i]),">25s"))
                else:
                    dr.append(self.ledbal1[i])
                    print(format(i,">22s"), format(str(self.ledbal1[i]),">14s"))

        for i in (self.Expenses):
            if i in result1:
                if self.ledbal1[i]>0:
                    dr.append(self.ledbal1[i])
                    print(format(i,">22s"), format(str(self.ledbal1[i]),">14s"))
                else:
                    cr.append(self.ledbal1[i])
                    print(format(i,">22s"), format(str(self.ledbal1[i]),">25s"))

        print("__________________________________________________________")
        print(format(sum(dr),">37"),format(sum(cr),">10"))
        print("__________________________________________________________")
    def AdGeneralEnteries(self):
        print("\n")
        print("--------------------------Adjusted General Enteries----------------------------------")
        print(self.read1)
        sumdr=self.read1['Debit'].sum()
        sumcr=self.read1['Credit'].sum()
        print("___________________________________________________________________________")
        print(format(sumdr,">63"),format(sumcr,">7"))
        print("___________________________________________________________________________")
    def AdAllLedger(self,AccountName):
        if AccountName=="Capital":
            result1 = list(self.read['Account Title and Explanation'])
            result2 = list(self.read1['Account Title and Explanation'])
            if AccountName in result1 and AccountName not in result2:
                Dr = self.read.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                Cr = self.read.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
                if AccountName in self.Expenses or AccountName in self.Drawings or AccountName in self.Revenue:
                    if Dr > Cr:
                        bal = (Dr) - (Cr)
                        self.ledbal[AccountName] = bal
                        bal1 = bal - bal
                    if Cr > Dr:
                        bal = (Dr) - (Cr)
                        self.ledbal[AccountName] = bal
                        bal1 = bal - bal
                else:
                    if Dr > Cr:
                        bal = (Dr) - (Cr)
                        self.ledbal[AccountName] = bal
                    if Cr > Dr:
                        bal = (Cr) - (Dr)
                        self.ledbal[AccountName] = bal
            elif AccountName in result1 and result2:
                Dr1 = self.read1.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                Cr1 = self.read1.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
                Dr = self.read.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                Cr = self.read.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
                if AccountName in self.Expenses or AccountName in self.Drawings or AccountName in self.Revenue:
                    if Dr > Cr:
                        bal = (Dr) - (Cr)
                        self.ledbal[AccountName] = bal
                        bal1 = bal - bal
                    if Cr > Dr:
                        bal = (Cr) - (Dr)
                        self.ledbal[AccountName] = bal
                        bal1 = bal - bal
                else:
                    if (Dr + Dr1) > (Cr + Cr1):
                        bal = (Dr + Dr1) - (Cr + Cr1)
                        self.ledbal[AccountName] = bal
                    if (Cr + Cr1) > (Dr + Dr1):
                        bal = (Cr + Cr1) - (Dr + Dr1)
                        self.ledbal[AccountName] = bal
                    else:
                        if AccountName in self.Expenses or AccountName in self.Drawings or AccountName in self.Revenue:
                            if Dr > Cr:
                                bal = (Dr) - (Cr)
                                self.ledbal[AccountName] = bal
                                bal1 = bal - bal
                            if Cr > Dr:
                                bal = (Cr) - (Dr)
                                self.ledbal[AccountName] = bal
                                bal1 = bal - bal
            else:
                Dr = self.read1.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                Cr =self.read1.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
                if AccountName in self.Expenses or AccountName in self.Drawings or AccountName in self.Revenue:
                    if Dr > Cr:
                        bal = (Dr) - (Cr)
                        self.ledbal[AccountName] = bal
                        bal1 = bal - bal
                    if Cr > Dr:
                        bal = (Cr) - (Dr)
                        self.ledbal[AccountName] = bal
                        bal1 = bal - bal
                else:
                    Dr1 = self.read1.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                    Cr1 = self.read1.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
                    if Dr1 > Cr1:
                        bal = (Dr1) - (Cr1)
                        self.ledbal[AccountName] = bal
                    if Cr1 > Dr1:
                        bal = (Cr1) - (Dr1)
                        self.ledbal[AccountName] = bal

        else:
            l = self.data.loc[self.data['Account Title and Explanation'] == AccountName].groupby(['Date', "Debit"]).sum().reset_index()
            result1 = list(self.read['Account Title and Explanation'])
            result2 = list(self.read1['Account Title and Explanation'])
            if AccountName in result1 and AccountName not in result2:
                print("----------------Ledger of ", AccountName, "------------------")
                print("_____________________________________________________________")
                Dr = self.read.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                Cr =self.read.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
                if AccountName in self.Expenses or AccountName in self.Drawings or AccountName in self.Revenue:
                    if Dr > Cr:
                        print(l)
                        bal = (Dr) - (Cr)
                        self.ledbal[AccountName] = bal
                        bal1 = bal - bal
                        print(format(("Bal:" + str(bal)), ">21"))
                        print("_____________________________________________________________")
                        print(format(bal1, ">16"), format(bal, ">9"))
                        print("_____________________________________________________________")
                        print(format(("Bal:" + str(bal1)), ">19"))
                        print("_____________________________________________________________")
                    if Cr > Dr:
                        print(l)
                        bal = (Dr) - (Cr)
                        self.ledbal[AccountName] = bal
                        bal1 = bal - bal
                        print(format(("Bal:" + str(bal)), ">21"))
                        print("_____________________________________________________________")
                        print(format(bal, ">16"), format(bal1, ">20"))
                        print("_____________________________________________________________")
                        print(format(("Bal:" + str(bal1)), ">16"))
                        print("_____________________________________________________________")
                else:
                    if Dr > Cr:
                        print(l)
                        bal = (Dr) - (Cr)
                        self.ledbal[AccountName] = bal
                        print(format(("Bal:" + str(bal)), ">23"))
                        print("_____________________________________________________________")
                    if Cr > Dr:
                        print(l)
                        bal = (Cr) - (Dr)
                        self.ledbal[AccountName] = bal
                        print(format(("Bal:" + str(bal)), ">29"))
                        print("_____________________________________________________________")
            elif AccountName in result1 and result2:
                Dr1 = self.read1.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                Cr1 = self.read1.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
                Dr = self.read.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                Cr = self.read.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]

                print("-------------------------- Ledger", AccountName,
                          "----------------------------------")
                print("_____________________________________________________________")
                if AccountName in self.Expenses or AccountName in self.Drawings or AccountName in self.Revenue:
                        if Dr > Cr:
                            print(l)
                            bal = (Dr+Dr1) - (Cr+Cr1)
                            self.ledbal[AccountName] = bal
                            bal1 = bal - bal
                            print(format(("Bal:" + str(bal)), ">21"))
                            print("_____________________________________________________________")
                            print(format(bal1, ">16"), format(bal, ">20"))
                            print("_____________________________________________________________")
                            print(format(("Bal:" + str(bal1)), ">16"))
                            print("_____________________________________________________________")

                        if Cr > Dr:
                            print(l)
                            bal = (Cr+Cr1) - (Dr+Dr1)
                            self.ledbal[AccountName] = bal
                            bal1 = bal - bal
                            print(format(("Bal:" + str(bal)), ">27"))
                            print("_____________________________________________________________")
                            print(format(bal, ">16"), format(bal1, ">20"))
                            print("_____________________________________________________________")
                            print(format(("Bal:" + str(bal1)), ">16"))
                            print("_____________________________________________________________")
                else:
                        if (Dr + Dr1) > (Cr + Cr1):
                            print(l)
                            bal = (Dr + Dr1) - (Cr + Cr1)
                            self.ledbal[AccountName] = bal
                            print(format(("Bal:" + str(bal)), ">21"))
                            print("_____________________________________________________________")
                        if (Cr + Cr1) > (Dr + Dr1):
                            print(l)
                            bal = (Cr + Cr1) - (Dr + Dr1)
                            self.ledbal[AccountName] = bal
                            print(format(("Bal:" + str(bal)), ">27"))
                        else:
                            if AccountName in self.Expenses or AccountName in self.Drawings or AccountName in self.Revenue:
                                if Dr > Cr:
                                    print(l)
                                    bal = (Dr) - (Cr)
                                    self.ledbal[AccountName] = bal
                                    bal1 = bal - bal
                                    print(format(("Bal:" + str(bal)), ">21"))
                                    print("_____________________________________________________________")
                                    print(format(bal1, ">16"), format(bal, ">20"))
                                    print("_____________________________________________________________")
                                    print(format(("Bal:" + str(bal1)), ">16"))
                                    print("_____________________________________________________________")

                                if Cr > Dr:
                                    print(l)
                                    bal = (Cr) - (Dr)
                                    self.ledbal[AccountName] = bal
                                    bal1 = bal - bal
                                    print(format(("Bal:" + str(bal)), ">27"))
                                    print("_____________________________________________________________")
                                    print(format(bal, ">16"), format(bal1, ">20"))
                                    print("_____________________________________________________________")
                                    print(format(("Bal:" + str(bal1)), ">16"))
                                    print("_____________________________________________________________")
            else:
                Dr = self.read1.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                Cr =self.read1.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
                if AccountName in self.Expenses or AccountName in self.Drawings or AccountName in self.Revenue:
                        print("-------------------------- Ledger", AccountName,
                          "----------------------------------")
                        if Dr > Cr:
                            print(l)
                            bal = (Dr) - (Cr)
                            self.ledbal[AccountName] = bal
                            bal1 = bal - bal
                            print(format(("Bal:" + str(bal)), ">21"))
                            print("_____________________________________________________________")
                            print(format(bal1, ">16"), format(bal, ">20"))
                            print("_____________________________________________________________")
                            print(format(("Bal:" + str(bal1)), ">16"))
                            print("_____________________________________________________________")

                        if Cr > Dr:
                            print(l)
                            bal = (Cr) - (Dr)
                            self.ledbal[AccountName] = bal
                            bal1 = bal - bal
                            print(format(("Bal:" + str(bal)), ">27"))
                            print("_____________________________________________________________")
                            print(format(bal, ">16"), format(bal1, ">20"))
                            print("_____________________________________________________________")
                            print(format(("Bal:" + str(bal1)), ">16"))
                            print("_____________________________________________________________")
                else:
                        print("-------------------------- Ledger", AccountName,
                              "----------------------------------")
                        Dr1 =self.read1.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                        Cr1 =self.read1.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
                        if Dr1 > Cr1:
                            print(l)
                            bal = (Dr1) - (Cr1)
                            self.ledbal[AccountName] = bal
                            print(format(("Bal:" + str(bal)), ">16"))
                            print("_____________________________________________________________")

                        if Cr1 > Dr1:
                            print(l)
                            bal = (Cr1) - (Dr1)
                            self.ledbal[AccountName] = bal
                            print(format(("Bal:" + str(bal)), ">27"))
                            print("_____________________________________________________________")
    def AdLedger(self,AccountName):
        result1 = list(self.read['Account Title and Explanation'])
        result2 = list(self.read1['Account Title and Explanation'])
        if AccountName in result1 and AccountName not in result2:
            l = self.read.loc[self.read['Account Title and Explanation'] == AccountName].groupby(
                ['Date', "Debit"]).sum().reset_index()
            print("----------------Ledger of ", AccountName, "------------------")
            print("_____________________________________________________________")
            Dr = self.read.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
            Cr = self.read.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
            if Dr > Cr:
                print(l)
                bal = (Dr) - (Cr)
                self.ledbal[AccountName] = bal
                print(format(("Bal:" + str(bal)), ">23"))
                print("_____________________________________________________________")
            if Cr > Dr:
                print(l)
                bal = (Cr) - (Dr)
                self.ledbal[AccountName] = bal
                print(format(("Bal:" + str(bal)), ">29"))
                print("_____________________________________________________________")
        elif AccountName in result1 and result2:
            l = self.Adled.loc[self.Adled['Account Title and Explanation'] == AccountName].groupby(
                ['Date', "Debit"]).sum().reset_index()
            Dr1 = self.read1.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
            Cr1 = self.read1.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
            Dr = self.read.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
            Cr = self.read.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]

            print("-------------------------- Ledger", AccountName,
                  "----------------------------------")
            print("_____________________________________________________________")

            if (Dr + Dr1) > (Cr + Cr1):
                print(l)
                bal = (Dr + Dr1) - (Cr + Cr1)
                self.ledbal[AccountName] = bal
                print(format(("Bal:" + str(bal)), ">21"))
                print("_____________________________________________________________")
            if (Cr + Cr1) > (Dr + Dr1):
                print(l)
                bal = (Cr + Cr1) - (Dr + Dr1)
                self.ledbal[AccountName] = bal
                print(format(("Bal:" + str(bal)), ">27"))

        else:
            l = self.read1.loc[self.read1['Account Title and Explanation'] == AccountName].groupby(
                ['Date', "Debit"]).sum().reset_index()
            print("-------------------------- Ledger", AccountName,
                   "----------------------------------")
            Dr1 = self.read1.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
            Cr1 = self.read1.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
            if Dr1 > Cr1:
                print(l)
                bal = (Dr1) - (Cr1)
                self.ledbal[AccountName] = bal
                print(format(("Bal:" + str(bal)), ">16"))
                print("_____________________________________________________________")

            if Cr1 > Dr1:
                print(l)
                bal = (Cr1) - (Dr1)
                self.ledbal[AccountName] = bal
                print(format(("Bal:" + str(bal)), ">27"))
                print("_____________________________________________________________")

    def ATAdLedger(self, AccountName):
        result1 = list(self.read['Account Title and Explanation'])
        result2 = list(self.read1['Account Title and Explanation'])
        if AccountName in result1 and AccountName not in result2:
            Dr = self.read.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
            Cr = self.read.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
            if Dr > Cr:
                bal = (Dr) - (Cr)
                self.ledbal[AccountName] = bal
            if Cr > Dr:
                bal = (Cr) - (Dr)
                self.ledbal[AccountName] = bal
        elif AccountName in result1 and result2:
            Dr1 = self.read1.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
            Cr1 = self.read1.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
            Dr = self.read.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
            Cr = self.read.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
            if AccountName in result1:
                if (Dr + Dr1) > (Cr + Cr1):
                    bal = (Dr + Dr1) - (Cr + Cr1)
                    self.ledbal[AccountName] = bal

                if (Cr + Cr1) > (Dr + Dr1):
                    bal = (Cr + Cr1) - (Dr + Dr1)
                    self.ledbal[AccountName] = bal

        else:

            Dr1 = self.read1.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
            Cr1 = self.read1.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
            if Dr1 > Cr1:
                bal = (Dr1) - (Cr1)
                self.ledbal[AccountName] = bal

            if Cr1 > Dr1:
                bal = (Cr1) - (Dr1)
                self.ledbal[AccountName] = bal

    def LAccountNames(self):
        for col in self.sheet1.iter_cols(min_col=2, max_col=5, min_row=2, values_only=True):
            self.exceldata1.append(col)
        for col in self.sheet.iter_cols(min_col=2, max_col=5, min_row=2, values_only=True):
            self.exceldata.append(col)
        accnames = list(self.exceldata[1])

        accnames.extend(self.exceldata1[1])

        accnames = list(dict.fromkeys(accnames))
        for i in accnames:
            self.Uaccnames.append(i)
        accnames.pop(-1)
        return accnames

    def AdAllledger(self):
        listAccNames=self.LAccountNames()
        for i in listAccNames:
            self.AdAllLedger(i)

    def AdAllLedgers(self):
        listAccNames = self.LAccountNames()
        for i in listAccNames:
            self.AdLedger(i)

    def ATAdAllLedger(self):
        listAccNames = self.LAccountNames()
        for i in listAccNames:
            self.ATAdLedger(i)


    def AdTrialBalance(self):
        self.ATAdAllLedger()
        print("------------------ Adjusted Trial  Balance------------------------")
        print("-------------------Company Name-------------------------")
        print("----------------For the Month (date)---------------------")
        print("_______________________________________________________")
        print("          Accounts              Debit      Credit")
        print("________________________________________________________")
        cr = []
        dr = []
        result1 = list(self.read['Account Title and Explanation'])
        result2 = list(self.read1['Account Title and Explanation'])

        for i in (self.Assets):
            if i in result1 and result2 or i in  result2:
                if self.ledbal[i] > 0:
                    dr.append(self.ledbal[i])
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">14s"))
                else:
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">20s"))
                    cr.append(self.ledbal[i])

        for i in (self.Liablities):
            if i in result1 and result2 or i in result2:
                if self.ledbal[i] > 0:
                    cr.append(self.ledbal[i])
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">25s"))
                else:
                    dr.append(self.ledbal[i])
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">14s"))

        for i in (self.Capital):
            if i in result1 and result2 or  i in result2:
                if self.ledbal[i] > 0:
                    cr.append(self.ledbal[i])
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">25s"))
                else:
                    dr.append(self.ledbal[i])
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">14s"))

        for i in (self.Drawings):
            if i in result1 and result2 or i in result2:
                if self.ledbal[i] > 0:
                    dr.append(self.ledbal[i])
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">14s"))
                else:
                    cr.append(self.ledbal[i])
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">25s"))

        for i in (self.Revenue):
            if i in result1 and result2 or i in  result2:
                if self.ledbal[i] > 0:
                    cr.append(self.ledbal[i])
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">25s"))
                else:
                    dr.append(self.ledbal[i])
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">14s"))

        for i in (self.Expenses):
            if i in result1 and result2 or i in  result2:
                if self.ledbal[i] > 0:
                    dr.append(self.ledbal[i])
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">14s"))
                else:
                    cr.append(self.ledbal[i])
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">25s"))

        print("__________________________________________________________")
        print(format(sum(dr), ">37"), format(sum(cr), ">10"))
        print("__________________________________________________________")



    def IncomeStatent(self):
        print("\n")
        print("------------------Income Statement------------------------")
        print("-------------------Company Name-------------------------")
        print("----------------For the Month (date)---------------------")
        print("\n")
        print("Revenues")
        print("________________________________________________________")
        revenue=[]
        expense=[]
        for i in (self.Revenue):
            if i in self.Uaccnames:
                if self.ledbal[i] > 0:
                    revenue.append(self.ledbal[i])
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">25s"))
        print("__________________________________________________________")
        print("Expenses")
        print("__________________________________________________________")
        for i in (self.Expenses):
            if i in self.Uaccnames:
                if self.ledbal[i]>0:
                    expense.append(self.ledbal[i])
                    print(format(i,">22s"), format(str(self.ledbal[i]),">14s"))

        print("__________________________________________________________")
        print(format("Total Expense",">26s"),format("(",">16"),sum(expense),")")

        print("__________________________________________________________")
        if sum(revenue)>sum(expense):
            NetIncome = sum(revenue) - sum(expense)
            self.NetIncome[0]=(NetIncome)
            print("Net Income",format(NetIncome,">37"))
        else:
            NetIncome = abs(sum(revenue) - sum(expense))
            self.NetIncome[0]=(NetIncome)
            print("Net Income", format("(",">33"),abs(NetIncome),")")


    def EquityStatement(self):
        print("--------------------Company Name-----------------------------")
        print("------------------Owner's Equity Statement------------------------")
        print("----------------For the Month (date)------------------------------")
        print("\n")
        if "Investment" in self.Uaccnames:
            Investment=[]
            if self.ledbal["Investment"]>0:
                Investment.append(self.ledbal["Investment"])
            capital = []
            date = self.read["Date"]
            for i in (self.Capital):
                if i in self.Uaccnames:
                    if i=="Investment":
                       pass
                    else:
                        if self.ledbal[i] >=0:
                            capital.append(self.ledbal[i])
                            print("Owner's Capital,", date[0], format(sum(capital), ">22"))
            print(format("Add: Investment", ">20s"), format(sum(Investment), ">25"))
            Add = sum(Investment ,self.NetIncome[0])
            print(format("Net Income", ">20s"), format(self.NetIncome[0], ">25"), format(Add, ">12"))
            Add1 = sum(capital) + (Add)
            # Add1 = 200
            print("________________________________________________________________________")
            print(format(Add1, ">59"))
            drawings = 0
            #for i in (self.Drawings):
            if self.Drawings[0] in self.Uaccnames:
                drawings = self.ledbal[self.Drawings[0]]
                print(format("less: Drawings", ">18s"), format(drawings, ">39"))
            else:
                print(format("less: Drawings", ">18s"), format(drawings, ">39"))
            date1 = self.read1["Date"]
            Sub = Add1 - drawings
            self.Ownersequity[0]=(Sub)
            print("________________________________________________________________________")
            if Sub >= 0:
                print("Owner's Capital,", date1[0], format(Sub, ">22"))
            else:
                Sub = abs(Sub)
                print("Owner's Capital,", date1[0], format("(", ">18"), Sub, ")")
        else:
            Investment =0
            capital = []
            date = self.read["Date"]
            for i in (self.Capital):
                if i in self.Uaccnames:
                    if self.ledbal[i] >= 0:
                        capital.append(self.ledbal[i])
                        print("Owner's Capital,", date[0], format(sum(capital), ">22"))
            print(format("Add: Investment", ">20s"), format((Investment), ">25"))
            Add = (Investment+self.NetIncome[0])
            print(format("Net Income", ">20s"), format(self.NetIncome[0], ">25"), format(Add, ">12"))
            Add1 = sum(capital) + (Add)
            # Add1 = 200
            print("________________________________________________________________________")
            print(format(Add1, ">59"))
            drawings = 0
            # for i in (self.Drawings):
            if self.Drawings[0] in self.Uaccnames:
                drawings = self.ledbal[self.Drawings[0]]
                print(format("less: Drawings", ">18s"), format(drawings, ">39"))
            else:
                print(format("less: Drawings", ">18s"), format(drawings, ">39"))
            date1 = self.read1["Date"]
            Sub = Add1 - drawings
            self.Ownersequity[0] = (Sub)
            print("________________________________________________________________________")
            if Sub >= 0:
                print("Owner's Capital,", date1[0], format(Sub, ">22"))
            else:
                Sub = abs(Sub)
                print("Owner's Capital,", date1[0], format("(", ">18"), Sub, ")")

    def BalanceSheet(self):
        print("--------------------Company Name-----------------------------")
        print("------------------Balance Sheet------------------------")
        print("----------------For the Month (date)------------------------------")
        print("\n")
        Assets=[]
        liablities=[ ]
        skip=[]
        contrabal=[]
        print("--------------------------Assets--------------------------------")
        for i in (self.Assets):
            if i in self.Uaccnames:
                if self.ledbal[i] >= 0:
                    Assets.append(self.ledbal[i])
                    print(i, format(self.ledbal[i], ">22"))

            if ("Accumulated Depreciation " + str(i)) in self.Uaccnames:
                if self.ledbal["Accumulated Depreciation "+i] >= 0:
                    Assets.append(self.ledbal["Accumulated Depreciation "+i])
                    contrabal.append(self.ledbal["Accumulated Depreciation " + i])
                    skip.append("Accumulated Depreciation "+i)
                    a = (self.ledbal[i]) - self.ledbal["Accumulated Depreciation " + i]
                    print(" Less: Accumulated Depreciation ",i,format("(",">10"), self.ledbal["Accumulated Depreciation "+i],")", format(a,">10"))
        print("____________________________________________________________________________________")
        print(format("Total Assets",">25"), format((sum(Assets) - sum(2*contrabal)),">43"))
        print("____________________________________________________________________________________")
        print("\n")
        print("-----------------------Liabilities And Owners Equity---------------------")
        print("Liabilities")
        for i in self.Liablities:
            if i in  skip:
                self.Liablities.remove(i)

        for i in self.Liablities:
            if i in self.Uaccnames:
                if self.ledbal[i] >= 0:
                    liablities.append(self.ledbal[i])
                    print(format(i,">25"), format(self.ledbal[i], ">22"))

        print("____________________________________________________________________________________")
        print(format("Total Liabilities", ">30"), format(sum(liablities), ">38"))

        self.Liablities.extend(skip)


        if "Capital" in self.Uaccnames:
            print("Owner's Equity")
            liablities.append(self.Ownersequity[0])
            print(format("Owner's Capital",">20"),format(self.Ownersequity[0], ">48"))

        print("____________________________________________________________________________________")
        print(format("Total liabilities and Owners Equity",">48"),format(sum(liablities), ">20"))



    def ClosingEntry(self):
        print("--------------------------Closing Enteries----------------------------------")
        print("\n")
        print("Revenues")

        #Storing sum of all revenue accounts
        Sum_ledbal_rev = 0

        # Storing sum of all revenue accounts
        Sum_ledbal_exp = 0

        # Adding balance Printing all revenue accounts
        for i in (self.Revenue):
            if i in self.Uaccnames:
                if self.ledbal[i] > 0:
                    print(format(i,">20") , format(self.ledbal[i],">20"))
                    Sum_ledbal_rev += self.ledbal[i]

        # Transferring sum of all revenue accounts to Income Summary
        if Sum_ledbal_rev > 0:
            print(format(self.IncomeSummary[0], ">25"), format(Sum_ledbal_rev,">25") )

        # Adding balance of all expense accounts
        for i in (self.Expenses):
            if i in self.Uaccnames:
                if self.ledbal[i] > 0:
                    Sum_ledbal_exp += self.ledbal[i]

        # Transferring sum of all expense accounts to Income Summary
        if Sum_ledbal_exp > 0:
            print(format(self.IncomeSummary[0], ">20"), format(Sum_ledbal_exp, ">20"))

            #Printing all expense accounts
            for i in (self.Expenses):
                if i in self.Uaccnames:
                    if self.ledbal[i] > 0:
                        print(format(i, ">25"), format(self.ledbal[i], ">25"))

        #Transferring Income Summary to Capital
        if Sum_ledbal_rev >= Sum_ledbal_exp:
            IS_Bal = Sum_ledbal_rev - Sum_ledbal_exp
            if IS_Bal == self.NetIncome[0]:
                if IS_Bal > 0:
                    print(format(self.IncomeSummary[0], ">20"), format(IS_Bal, ">20"))
                    print(format(self.Capital[0], ">25"), format(IS_Bal, ">25"))

                elif IS_Bal < 0:
                    print(format(self.Capital[0], ">20"), format(abs(IS_Bal), ">20"))
                    print(format(self.IncomeSummary[0], ">25"), format(abs(IS_Bal), ">25"))

                else:
                    pass
            else:
                print("Income Summary ledger is not equal to Net Income")
                if IS_Bal > self.NetIncome[0]:
                    print("Income Summary ledger is greater than Net Income")
                else:
                    print("Income Summary ledger is less than Net Income")

        # Transferring Drawings to Capital
        Sum_ledbal_draw = 0
        for i in (self.Drawings):
            if i in self.Uaccnames:
                if self.ledbal[i] > 0:
                    Sum_ledbal_draw +=self.ledbal[i]
        for i in (self.Drawings):
            if i in self.Uaccnames:
                print(format(self.Capital[0], ">14"), format (Sum_ledbal_draw, ">28"))
                print(format(i, ">25"), format(Sum_ledbal_draw, ">25"))
    def PCClosingLedger(self, AccountName):
        result1 = list(self.read['Account Title and Explanation'])
        result2 = list(self.read1['Account Title and Explanation'])
        if AccountName in result1 and AccountName in result2:
            if AccountName in self.Revenue:
                Dr1 = self.read1.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                Cr1 = self.read1.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
                Dr = self.read.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                Cr = self.read.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]

                if (Dr + Dr1) > (Cr + Cr1):
                    bal = (Dr + Dr1) - (Cr + Cr1)
                    self.ledbal1[AccountName] = bal
                    bal1 = bal - bal

                if (Cr + Cr1) > (Dr + Dr1):
                    bal = (Cr + Cr1) - (Dr + Dr1)
                    self.ledbal1[AccountName] = bal
                    bal1 = bal - bal


            if AccountName in self.Expenses:
                Dr1 = self.read1.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                Cr1 = self.read1.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
                Dr = self.read.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                Cr = self.read.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]

                if (Dr + Dr1) > (Cr + Cr1):
                    bal = (Dr + Dr1) - (Cr + Cr1)
                    self.ledbal1[AccountName] = bal
                    bal1 = bal - bal


                if (Cr + Cr1) > (Dr + Dr1):
                    bal = (Cr + Cr1) - (Dr + Dr1)
                    self.ledbal1[AccountName] = bal
                    bal1 = bal - bal


            if AccountName in self.Drawings:
                Dr1 = self.read1.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                Cr1 = self.read1.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
                Dr = self.read.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                Cr = self.read.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]

                if (Dr + Dr1) > (Cr + Cr1):
                    bal = (Dr + Dr1) - (Cr + Cr1)
                    self.ledbal1[AccountName] = bal
                    bal1 = bal - bal

                if (Cr + Cr1) > (Dr + Dr1):
                    bal = (Cr + Cr1) - (Dr + Dr1)
                    self.ledbal1[AccountName] = bal
                    bal1 = bal - bal




        elif AccountName in result1 or result2:
            if AccountName in self.Revenue:
                if AccountName in result1:
                    Dr = self.read.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                    Cr = self.read.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]

                    if Dr > Cr:
                        bal = (Dr) - (Cr)
                        self.ledbal1[AccountName] = bal
                        bal1 = bal - bal

                    if Cr > Dr:
                        bal = (Cr) - (Dr)
                        self.ledbal1[AccountName] = bal
                        bal1 = bal - bal


                else:
                    Dr1 = self.read1.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                    Cr1 = self.read1.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]

                    if Dr1 > Cr1:
                        bal = (Dr1) - (Cr1)
                        self.ledbal1[AccountName] = bal
                        bal1 = bal - bal


                    if Cr1 > Dr1:
                        bal = (Cr1) - (Dr1)
                        self.ledbal1[AccountName] = bal
                        bal1 = bal - bal

            if AccountName in self.Expenses:
                if AccountName in result1:
                    Dr = self.read.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                    Cr = self.read.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]

                    if Dr > Cr:
                        bal = (Dr) - (Cr)
                        self.ledbal1[AccountName] = bal
                        bal1 = bal - bal


                    if Cr > Dr:
                        bal = (Cr) - (Dr)
                        self.ledbal1[AccountName] = bal
                        bal1 = bal - bal


                else:
                    Dr1 = self.read1.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                    Cr1 = self.read1.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]

                    if Dr1 > Cr1:
                        bal = (Dr1) - (Cr1)
                        self.ledbal1[AccountName] = bal
                        bal1 = bal - bal


                    if Cr1 > Dr1:
                        bal = (Cr1) - (Dr1)
                        self.ledbal1[AccountName] = bal
                        bal1 = bal - bal


            if AccountName in self.Drawings:
                if AccountName in result1:
                    Dr = self.read.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                    Cr = self.read.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]

                    if Dr > Cr:
                        bal = (Dr) - (Cr)
                        self.ledbal1[AccountName] = bal
                        bal1 = bal - bal


                    if Cr > Dr:
                        bal = (Cr) - (Dr)
                        bal1 = bal - bal

                else:
                    Dr1 = self.read1.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
                    Cr1 = self.read1.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]

                    if Dr1 > Cr1:
                        bal = (Dr1) - (Cr1)
                        self.ledbal1[AccountName] = bal
                        bal1 = bal - bal

                    if Cr1 > Dr1:
                        bal = (Cr1) - (Dr1)
                        self.ledbal1[AccountName] = bal
                        bal1 = bal - bal

    def ClsAccountNames(self):
        for col in self.sheet1.iter_cols(min_col=2, max_col=5, min_row=2, values_only=True):
            self.exceldata1.append(col)
        for col in self.sheet.iter_cols(min_col=2, max_col=5, min_row=2, values_only=True):
            self.exceldata.append(col)
        accnames = list(self.exceldata[1])

        accnames.extend(self.exceldata1[1])

        accnames = list(dict.fromkeys(accnames))
        for i in accnames:
            self.Uaccnames.append(i)
        accnames.pop(-1)
        return accnames


    def CClosingAllledger(self):
        listAccNames=self.ClsAccountNames()
        for i in listAccNames:
            self.PCClosingLedger(i)
    def ClosingAllledger(self):
        listAccNames=self.ClsAccountNames()
        for i in listAccNames:
            self.PCClosingLedger(i)
    def InSumLedger(self):
        Dr = 0
        Cr = 0
        Bal = 0
        Dr1 = 0
        Bal1 = 0
        Cr1 = 0
        result1 = list(self.read['Account Title and Explanation'])
        result2 = list(self.read1['Account Title and Explanation'])
        for i in self.Revenue:
            if i in result2:
                    if self.ledbal[i] > 0:
                        Cr += self.ledbal[i]
                        self.ledbal[i] = 0
        for i in self.Expenses:
            if i in result1 and result2:
                if self.ledbal[i] > 0:
                    Dr += self.ledbal[i]
                    self.ledbal[i] = 0
            else:
                if i in result1:
                    if self.ledbal[i] > 0:
                        Dr += self.ledbal[i]
                        self.ledbal[i] = 0

                if i in result2:
                    if self.ledbal[i] > 0:
                        Dr += self.ledbal[i]
                        self.ledbal[i] = 0
        if (Cr or Dr) > 0:
            if Cr > 0:
                self.data = self.data.append(
                    {'Date': date.today(), 'Account Code': "xxx", 'Account Title and Explanation': 'Income Summary',"Debit": 0,
                      "Credit":int(Cr)}, ignore_index=True)
            if Dr > 0:
                self.data = self.data.append(
                    {'Date': date.today(), 'Account Code': "xxx", 'Account Title and Explanation': 'Income Summary',
                     "Debit": int(Dr),"Credit":0}, ignore_index=True)
            if Cr >= Dr:
                Bal = Cr - Dr
                self.ISBal = Cr - Dr
                self.data = self.data.append(
                    {'Date': str(date.today()), 'Account Code': "xxx", 'Account Title and Explanation': 'Capital',
                     "Debit": 0, "Credit": int(Bal)}, ignore_index=True)


                for i in (self.Capital):
                    if i in self.Uaccnames:
                        if i == "Investment":
                            pass
                        else:
                            if self.ledbal[i] >= 0:

                                self.ledbal[i] = self.ledbal[i] + self.ISBal
            elif Dr > Cr:
                Bal = Dr - Cr
                self.data = self.data.append(
                    {'Date': date.today(), 'Account Code': "xxx", 'Account Title and Explanation': 'Capital',
                     "Debit": int(Bal), "Credit": 0}, ignore_index=True)

                for i in (self.Capital):
                    if i in self.Uaccnames:
                        if i == "Investment":
                            pass
                        else:
                            if self.ledbal[i] >= 0:
                                if self.ledbal[i] >= self.ISBal:
                                    self.ledbal[i] = self.ledbal[i] - self.ISBal
                                else:
                                    self.ledbal[i] = abs(self.ledbal[i] - self.ISBal)
        for i in self.Drawings:
            if i in result1 and result2:
                if self.ledbal[i] > 0:
                    Dr1 += self.ledbal[i]
                    self.ledbal[i] = 0
                    for j in (self.Capital):
                        if j in self.Uaccnames:
                            if j == "Investment":
                                pass
                            else:
                                if self.ledbal[j] >= 0:
                                    self.data = self.data.append(
                        {'Date':date.today(), 'Account Code': "xxx",
                         'Account Title and Explanation': 'Capital',
                         "Debit": Dr1, "Credit": 0}, ignore_index=True)
                                    self.ledbal[j] = self.ledbal[j] - Dr1
            else:
                if i in result1:
                    if self.ledbal[i] > 0:
                        Dr1 += self.ledbal[i]
                        self.ledbal[i] = 0
                        for j in (self.Capital):
                            if j in self.Uaccnames:
                                if j == "Investment":
                                    pass
                                else:
                                    if self.ledbal[j] >= 0:
                                        self.data = self.data.append(
                                                {'Date': date.today(), 'Account Code': "xxx",
                                                 'Account Title and Explanation': 'Capital',
                                                 "Debit": Dr1, "Credit": 0}, ignore_index=True)

                                        self.ledbal[j] = self.ledbal[j] - Dr1
                if i in result2:
                    if self.ledbal[i] > 0:
                        Dr1 += self.ledbal[i]
                        self.ledbal[i] = 0
                        for j in (self.Capital):
                            if j in self.Uaccnames:
                                if j == "Investment":
                                    pass
                                else:
                                    if self.ledbal[j] >= 0:
                                        self.data = self.data.append(
                                            {'Date': date.today(), 'Account Code': "xxx",
                                             'Account Title and Explanation': 'Capital',
                                             "Debit": Dr1, "Credit": 0}, ignore_index=True)
                                        self.ledbal[j] = self.ledbal[j] - Dr1
        for j in self.Capital:
            if j in result1 and result2:
                if j in self.Uaccnames:
                    if j == "Investment":
                        if self.ledbal[j] >= 0:
                            self.data = self.data.append(
                                {'Date': date.today(), 'Account Code': "xxx",
                                 'Account Title and Explanation': 'Capital',
                                 "Debit": 0, "Credit": int(self.ledbal[j])}, ignore_index=True)
                            Cr1 += self.ledbal[j]
                            self.ledbal[j] = 0
                            self.ledbal["Capital"] = self.ledbal["Capital"] + Cr1

            if j in result1:
                if j in self.Uaccnames:
                    if j == "Investment":
                        if self.ledbal[j] >= 0:
                            self.data = self.data.append(
                                {'Date': '2019-09-10', 'Account Code': "xxx",
                                 'Account Title and Explanation': 'Capital',
                                 "Debit": 0, "Credit": int(self.ledbal[j])}, ignore_index=True)
                            Cr1 += self.ledbal[j]

                            self.ledbal[j] = 0
                            self.ledbal["Capital"] = self.ledbal["Capital"] + Cr1
            if j in result2:
                if j in self.Uaccnames:
                    if j == "Investment":
                        if self.ledbal[j] >= 0:
                            self.data = self.data.append(
                                {'Date': date.today(), 'Account Code': "xxx",
                                 'Account Title and Explanation': 'Capital',
                                 "Debit": 0, "Credit": int(self.ledbal[j])}, ignore_index=True)
                            Cr1 += self.ledbal[j]

                            self.ledbal[j] = 0
                            self.ledbal["Capital"] = self.ledbal["Capital"] + Cr1


        self.LedgerForCI("Capital")
        self.LedgerForCI("Income Summary")
    def ClsTrialBalance(self):
        self.CClosingAllledger()
        print("------------------ Closing Trial  Balance------------------------")
        print("-------------------Company Name-------------------------")
        print("----------------For the Month (date)---------------------")
        print("_______________________________________________________")
        print("          Accounts              Debit      Credit")
        print("________________________________________________________")
        cr = []
        dr = []

        result1 = list(self.read['Account Title and Explanation'])
        result2 = list(self.read1['Account Title and Explanation'])
        for i in (self.Assets):
            if i in result1 and result2 or i in result2:
                if self.ledbal[i] > 0:
                    dr.append(self.ledbal[i])
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">14s"))
                else:
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">20s"))
                    cr.append(self.ledbal[i])
        for i in (self.Liablities):
            if i in result1 and result2 or i in result2:
                if self.ledbal[i] > 0:
                    cr.append(self.ledbal[i])
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">25s"))
                else:
                    dr.append(self.ledbal[i])
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">14s"))
        for i in (self.Capital):
            if i in result1 and result2 or i in result2:
                if self.ledbal[i] > 0:
                    cr.append(self.ledbal[i])
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">25s"))
                else:
                    dr.append(self.ledbal[i])
                    print(format(i, ">22s"), format(str(self.ledbal[i]), ">14s"))



        print("__________________________________________________________")
        print(format(sum(dr), ">37"), format(sum(cr), ">10"))
        print("__________________________________________________________")
    def LedgerForCI(self,AccountName):
        if AccountName=="Income Summary":
            print("\n")
            print("----------------Ledger of ", AccountName, "------------------")
            print("_____________________________________________________________")
            l = self.data.loc[self.data['Account Title and Explanation'] == AccountName].groupby(
                ['Date', "Debit"]).sum().reset_index()
            Dr = self.data.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
            Cr = self.data.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
            if Dr > Cr:
                print(l)
                bal = (Dr) - (Cr)
                self.ledbal[AccountName] = bal
                bal1 = bal - bal
                print(format(("Bal:" + str(bal)), ">19"))
                print("_____________________________________________________________")
                print(format(bal1, ">19"), format(bal, ">5"))
                print("_____________________________________________________________")
                print(format(("Bal:" + str(bal1)), ">19"))
                print("_____________________________________________________________")
            if Cr > Dr:
                print(l)
                bal = (Cr) - (Dr)
                self.ledbal[AccountName] = bal
                bal1 = bal - bal
                print(format(("Bal:" + str(bal)), ">30"))
                print("_____________________________________________________________")
                print(format(bal, ">19"), format(bal1, ">5"))
                print("_____________________________________________________________")
                print(format(("Bal:" + str(bal1)), ">19"))
                print("_____________________________________________________________")
        else:
            print("\n")
            print("----------------Ledger of ", AccountName, "------------------")
            print("_____________________________________________________________")
            l = self.data.loc[self.data['Account Title and Explanation'] == AccountName].groupby(
                ['Date', "Debit"]).sum().reset_index()
            Dr = self.data.groupby('Account Title and Explanation')['Debit'].sum()[AccountName]
            Cr = self.data.groupby('Account Title and Explanation')['Credit'].sum()[AccountName]
            if Dr > Cr:
                print(l)
                bal = (Dr) - (Cr)
                self.ledbal1[AccountName] = bal
                print("_____________________________________________________________")
                print(format(("Bal:" + str(bal)), ">21"))
                print("_____________________________________________________________")
            if Cr > Dr:
                print(l)
                bal = (Cr) - (Dr)
                self.ledbal1[AccountName] = bal

                print("_____________________________________________________________")
                print(format(("Bal:" + str(bal)), ">30"))
                print("_____________________________________________________________")

    def chk(self):
        for v in self.data["Credit"]:
            print(v)
obj=AccountingCycle("testfile1.xlsx","adjusted.xlsx")
obj.Data()
print(obj.GeneralEnteries())